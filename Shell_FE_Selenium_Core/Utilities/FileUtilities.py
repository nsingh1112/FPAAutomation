import csv
import os
import openpyxl
import json
import xmltodict
from Shell_FE_Selenium_Core.Utilities.LoggingUtilities import LoggingUtilities


class FileUtilities:
    """FileUtilities class contains reusable methods to read / write data into files (i.e.) Excel, Json etc."""

    current_working_directory = os.path.dirname(os.getcwd())
    test_data = current_working_directory + "/Shell_FE_Behave_Tests/TestData/"
    logobj = LoggingUtilities()
    log = logobj.logger()

    @staticmethod
    def read_excel_by_row_name(file_name, sheet_name, row_name):
        """Reads the data present in excel file based on row name and returns a dictionary of the values."""
        dictionary_values = dict()
        work_book = openpyxl.load_workbook(FileUtilities.test_data + file_name)
        work_sheet = work_book[sheet_name]
        max_row = work_sheet.max_row
        max_col = work_sheet.max_column
        row_name_available = False
        for x in range(1, max_row + 1):
            if work_sheet.cell(row=x, column=1).value == row_name:
                row_name_available = True
                for y in range(2, max_col + 1):
                    dictionary_values[work_sheet.cell(row=1, column=y).value] = work_sheet.cell(row=x, column=y).value
        if row_name_available is not True:
            FileUtilities.log.error("The row {0} is not available in the work sheet.".format(row_name))
            raise Exception("The row {0} is not available in the work sheet.".format(row_name))
        FileUtilities.log.info("The data from excel sheet with row {0} has been converted into dictionary.".format(row_name))
        return dictionary_values

    @staticmethod
    def read_cell_value_from_excel(file_name, sheet_name, cell_name):
        """Reads the data present in excel based on cell name and returns the single cell value."""
        work_book = openpyxl.load_workbook(FileUtilities.test_data + file_name)
        work_sheet = work_book[sheet_name]
        FileUtilities.log.info(
            "The value from excel sheet with cell {0} has been returned.".format(cell_name))
        return work_sheet[cell_name].value

    @staticmethod
    def read_json_file_as_dictionary(file_name):
        """Reads the data present in a json file and returns a dictionary of the values."""
        try:
            with open(FileUtilities.test_data + file_name, 'r') as json_file:
                dictionary_values = json.load(json_file)
                FileUtilities.log.info("The data from json file {0} has been converted into dictionary.".format(file_name))
                return dictionary_values
        except FileNotFoundError as f_err:
            FileUtilities.log.error("Unable to locate Json file {0}. Exception {1}".format(file_name, f_err))
        except Exception as err:
            FileUtilities.log.error("Unable to load Json file {0}. Exception {1}".format(file_name, err))

    @staticmethod
    def read_json_string_as_dictionary(json_string):
        """Reads the data present in a json string and returns a dictionary of the values."""
        if isinstance(json_string, str) is False:
            FileUtilities.log.error("Only a string value should be passed to the method read_json_string_as_dictionary(json_string).")
            raise ValueError("Attribute should be a string value!!")
        dictionary_values = json.loads(json_string)
        FileUtilities.log.info("The json string {0} has been converted into dictionary.".format(json_string))
        return dictionary_values

    @staticmethod
    def write_into_json(dict_name, file_name):
        """Converts the dictionary into a Json file. If no file available with the name then
        new file will be created else the file would be overwritten."""
        try:
            with open(FileUtilities.test_data + file_name, 'w') as json_file:
                json.dump(dict_name, json_file, indent=2)
                FileUtilities.log.info("File created with contents. Filename: {0}.".format(file_name))
        except TypeError as terr:
            FileUtilities.log.error(
                "Key that does not belong to a basic (i.e.) int / float / string / bool / None type is present. Exception: {0}".format(
                    terr))
        except Exception as err:
            FileUtilities.log.error("Unable to write into json file. Exception: {0}.".format(err))

    @staticmethod
    def convert_dictionary_to_json_string(dict_name):
        """Converts the dictionary into a json string."""
        json_string = ""
        try:
            json_string = json.dumps(dict_name, indent=2)
            FileUtilities.log.info("Converted the dictionary {0} to a json string.".format(dict_name))
            return json_string
        except TypeError as terr:
            FileUtilities.log.error(
                "Key that does not belong to a basic (i.e.) int / float / string / bool / None type is present. Exception: {0}".format(
                    terr))
        except Exception as err:
            FileUtilities.log.error("Unable to convert dictionary to Json string. Exception: {0}".format(err))

    @staticmethod
    def write_into_existing_excel_file(dict_name, file_name, sheet_name, row_name):
        """Writes data from dictionary into existing Excel based on the sheet name and row name.
        The row headers should match the dictionary keys.
        """
        work_book = openpyxl.load_workbook(FileUtilities.test_data + file_name)
        work_sheet = work_book[sheet_name]
        max_row = work_sheet.max_row
        max_col = work_sheet.max_column
        row_name_available = False
        for x in range(1, max_row + 1):
            if work_sheet.cell(row=x, column=1).value == row_name:
                row_name_available = True
                for y in range(2, max_col + 1):
                    work_sheet.cell(row=x, column=y).value = dict_name[work_sheet.cell(row=1, column=y).value]
        if row_name_available is not True:
            FileUtilities.log.error("The row {0} is not available in the work sheet.".format(row_name))
            raise Exception("The row {0} is not available in the work sheet.".format(row_name))
        work_book.save(FileUtilities.test_data + file_name)
        FileUtilities.log.info("Updated the excel file {0} with the contents from dictionary {1}.".format(file_name, dict_name))

    @staticmethod
    def write_into_new_excel_file(dict_name, file_name, sheet_name):
        """Writes data from dictionary into a new Excel based on the sheet name and row name.
        If file exists, the file contents would be overwritten.
        """
        work_book = openpyxl.Workbook()
        sheet = work_book.active
        sheet.title = sheet_name
        max_col = len(dict_name)
        index = 1
        for key in dict_name.keys():
            sheet.cell(row=1, column=index).value = key
            sheet.cell(row=2, column=index).value = dict_name[key]
            index += 1
        work_book.save(FileUtilities.test_data + file_name)
        FileUtilities.log.info("Wrote the contents of dictionary {0} into the excel file {1}".format(dict_name, file_name))

    @staticmethod
    def write_into_excel_via_cell_name(file_name, sheet_name, cell_name, value):
        """Writes the parameterized value into the specified cell name."""
        work_book = openpyxl.load_workbook(FileUtilities.test_data + file_name)
        work_sheet = work_book[sheet_name]
        work_sheet[cell_name].value = value
        work_book.save(FileUtilities.test_data + file_name)
        FileUtilities.log.info("Wrote the value {0} into the file {1}".format(value, file_name))

    @staticmethod
    def read_csv_by_row_name(file_name, row_header, row_name):
        """Reads the contents of a csv file and retrieves values as dictionary based on row name"""
        dictionary_csv = dict()
        try:
            with open(FileUtilities.test_data + file_name, 'r') as csv_file:
                csv_json = csv.DictReader(csv_file)
                row_name_available = False
                for row in csv_json:
                    if row[row_header] == row_name:
                        dictionary_csv = row
                        row_name_available = True
                        FileUtilities.log.info(
                            "The data from csv file {0} has been converted into dictionary.".format(file_name))
                        break
                if row_name_available is True:
                    del dictionary_csv[row_header]
                if row_name_available is False:
                    FileUtilities.log.warning(
                        "The row with row name {0} is not present in the csv file {1}.".format(row_name, file_name))
            return dictionary_csv
        except FileNotFoundError as f_err:
            FileUtilities.log.error("Unable to locate CSV file {0}. Exception {1}".format(file_name, f_err))
        except Exception as err:
            FileUtilities.log.error("Unable to parse CSV file {0}. Exception {1}".format(file_name, err))

    @staticmethod
    def read_xml(file_name, root_tag, child_tag=None, child_tag_index=None):
        """Reads the contents of a xml file and retrieves required values as dictionary.
            :Args:
                - file_name - The name of the XML file present in the TestData folder
                - root_tag - The root tag of the XML document
                - child_tag - The child tag whose values needs to be retrieved. It should
                be the immediate child of the root tag. This carries a default value of None.
                - child_tag_index - The index / position of the child tag (if the root tag contains
                 more than one child tag with same name). This carries a default value of None.
        """
        dictionary_xml = dict()
        try:
            with open(FileUtilities.test_data + file_name, 'r', encoding='utf-8') as xml_file:
                xml = xml_file.read()
                ordered_dictionary = xmltodict.parse(xml)
                if child_tag is None:
                    dictionary_xml = dict(ordered_dictionary[root_tag])
                else:
                    if child_tag_index is not None:
                        root_dict = dict(ordered_dictionary[root_tag])
                        dictionary_xml = dict(root_dict[child_tag][child_tag_index])
                    if child_tag_index is None:
                        root_dict = dict(ordered_dictionary[root_tag])
                        dictionary_xml = dict(root_dict[child_tag])
                return dictionary_xml
        except FileNotFoundError as f_err:
            FileUtilities.log.error("Unable to locate XML file {0}. Exception {1}".format(file_name, f_err))
        except Exception as err:
            FileUtilities.log.error("Unable to parse XML file {0}. Exception {1}".format(file_name, err))

    @staticmethod
    def search_word_from_pdf(file_name, word, pagenum=0, exact_match = False, substring = True):
        """
        Reads the pdf and seraches for word.
        :param file_name: provide the file name
        :param word: the word to be searched
        :param pagenum: on which page the pdf needs to be searched
        :param exact_match: if the exact word need to be find with case sensitivity
        :param substring: to find the subtring in a string
        :return: True or False if the word is found in pdf file based on exact match filter
        The count feature of exact word can also be used if required
        """
        try:
            pdf_file = open(FileUtilities.test_data + file_name, "rb")
            read_pdf = PyPDF2.PdfFileReader(pdf_file)
            getpage = read_pdf.getPage(pagenum)
            pdfData = getpage.extractText()

            pdfdatalist = pdfData.split()
            if exact_match:
                word_count = 0
                for i in range(len(pdfdatalist)):
                    ResSearch = re.fullmatch(word, pdfdatalist[i])
                    if ResSearch is not None:
                        word_count = word_count+1
                if word_count == 0:
                    FileUtilities.log.error("Word {0} not found in file {1}".format(word, file_name))
                    return False
                else:
                    FileUtilities.log.info("Word {0} found in file {1}".format(word, file_name))
                    return True

            elif substring == True:
                word_count = 0
                for i in range(len(pdfdatalist)):
                    ResSearch = re.search(word, pdfData)
                    if ResSearch is not None:
                        word_count = word_count+1
                if word_count == 0:
                    FileUtilities.log.error("Word {0} not found in file {1}".format(word, file_name))
                    return False
                else:
                    FileUtilities.log.info("Word {0} found in file {1}".format(word, file_name))
                    return True
            else:
                ResSearch = re.match(word, pdfData)
                if ResSearch is not None:
                    FileUtilities.log.info("Word {0} found in file {1}".format(word, file_name))
                    return True
                else:
                    FileUtilities.log.error("Word {0} not found in file {1}".format(word, file_name))
                    return False
        # Unable to open the file
        except IOError as ex:
            FileUtilities.log.error('IOError thrown. Details: ' + ' %s' % ex)
            return False
        # Any other exceptions
        except Exception as ex:
            FileUtilities.log.error('Unexpected error. Details:' + ' %s' % ex)
            return False

    @staticmethod
    def get_number_of_pages(file_name):
        """Return the number of pages from pdf.
            This function can also be used in test cases and used to run
            the loop for all the pages after getting the page count
        """
        try:
            pdf_file = open(FileUtilities.test_data + file_name, "rb")
            read_pdf = PyPDF2.PdfFileReader(pdf_file)
            number_of_pages = read_pdf.getNumPages()
            FileUtilities.log.info("Return the number of pages")
            return number_of_pages
        # Unable to open the file
        except IOError as ex:
            FileUtilities.log.error('IOError thrown. Details: ' + ' %s' % ex)
        # Any other exceptions
        except Exception as ex:
            FileUtilities.log.error('Unexpected error. Details:' + ' %s' % ex)
